from openpyxl import load_workbook
from openpyxl import Workbook
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

def generate_sentences_from_excel(file_path, output_path, batch_size=1000):
    """
    Efficiently process an Excel file with a large number of rows and save sentiment analysis to a new file.

    Args:
        file_path (str): Path to the Excel file.
        output_path (str): Path to save the output Excel file.
        batch_size (int): Number of rows to process in each batch.
    """
    try:
        # Load workbook and the active sheet
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active

        # Find the 'Content' column
        header = next(sheet.iter_rows(values_only=True))
        if 'Content' not in header:
            raise ValueError("The Excel file must contain a 'Content' column.")
        content_index = header.index('Content')

        # Initialize sentiment analyzer
        analyzer = SentimentIntensityAnalyzer()

        # Create a new workbook for output
        output_wb = Workbook()
        output_sheet = output_wb.active
        output_sheet.title = "Sentiment Analysis"

        # Write headers to the output file
        output_sheet.append(["Content", "Negative", "Neutral", "Positive", "Compound"])

        # Process rows in batches
        batch = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if i == 1:  # Skip header
                continue

            # Append row content to the current batch
            if row[content_index] is not None:
                batch.append(row[content_index])

            # If batch size is reached, process and reset batch
            if len(batch) >= batch_size:
                for sentence in batch:
                    vs = analyzer.polarity_scores(sentence)
                    output_sheet.append([sentence, vs['neg'], vs['neu'], vs['pos'], vs['compound']])
                batch = []

        # Process remaining rows in the batch
        if batch:
            for sentence in batch:
                vs = analyzer.polarity_scores(sentence)
                output_sheet.append([sentence, vs['neg'], vs['neu'], vs['pos'], vs['compound']])

        # Save the output workbook
        output_wb.save(output_path)
        print(f"Sentiment analysis results saved to {output_path}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    # Path to your input and output Excel files
    input_file_path = 'C:/Users/User/Desktop/KeyP/Liberal_Party_of_Australia.xlsx'
    output_file_path = 'C:/Users/User/Desktop/KeyP/Result_Liberal_Party_of_Australia.xlsx'

    # Generate sentiment analysis and save to output file
    generate_sentences_from_excel(input_file_path, output_file_path, batch_size=1000)



